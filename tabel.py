import openpyxl as opx
from random import randint as r
import datetime
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
import os
from random import shuffle as sh


def count_hour(mon, year, holiday, komandirovka, short_days_list):
    count_time = 0
    a = workday(1, mon, year, holiday, komandirovka)
    while True:
        try:
            nextworkday = next(a)
            if nextworkday[0]:
                count_time += 8
                if nextworkday[1] in short_days_list:
                    count_time -= 1
        except:
            break
    return count_time


def workday(d, m, y, holiday, komandirovka):
    komandirovka_list = []
    for i in komandirovka:
        for k in range(i[1], i[2] + 1):
            komandirovka_list.append(k)
    while True:
        try:
            if d in holiday or d in komandirovka_list:
                yield [False, d]
            else:
                yield [True if datetime.datetime(y, m, d).isoweekday() < 6 else False, d]
        except:
            break
        d += 1


def fill_tabel(all_time, workdaygen, short_days_list):
    all_stroki = []
    short_days = []
    work_list = []
    for ind, val in enumerate(all_time):
        all_time[ind] = int(val)
    revers_check = True
    while True:
        try:
            curent_day = []
            work_list.append(next(workdaygen))
            if work_list[-1][0]:
                revers_check = not revers_check
                if revers_check:
                    all_time.reverse()
                if work_list[-1][1] in short_days_list:
                    hourmax = 7
                else:
                    hourmax = 8
                for ind, top in enumerate(all_time):
                    if top > 0:
                        if ind == len(all_time) - 1:
                            hour = hourmax - sum(curent_day)
                            curent_day.append(hour)
                            all_time[ind] -= hour
                        elif sum(curent_day) < hourmax:
                            if sum(all_time[ind+1:]) > hourmax or (sum(curent_day) + sum(all_time[ind+1:]) > hourmax):
                                if top > hourmax:
                                    hour = r(1, hourmax - sum(curent_day))
                                elif top > hourmax - sum(curent_day):
                                    hour = hourmax - sum(curent_day)
                                else:
                                    hour = top
                                all_time[ind] -= hour
                                curent_day.append(hour)
                            elif sum(all_time[ind+1:]) < hourmax:
                                hour = hourmax - sum(curent_day) - sum(all_time[ind+1:])
                                all_time[ind] -= hour
                                curent_day.append(hour)
                        else:
                            curent_day.append(0)
                    else:
                        curent_day.append(0)
                while len(curent_day) != len(all_time):
                    curent_day.append(0)
                if revers_check:
                    curent_day.reverse()
                    all_time.reverse()
                if hourmax == 8:
                    all_stroki.append(curent_day)
                else:
                    short_days.append(curent_day)
            # else:
            #     all_stroki.append([' - ' for _ in range(len(all_time))])
        except:
            break
    sh(all_stroki)
    for i in work_list:
        if i[1] in short_days_list and short_days:
            all_stroki.insert(i[1] - 1, short_days.pop(0))
        elif not i[0]:
            all_stroki.insert(i[1] - 1, [' - ' for _ in range(len(all_time))])
    return all_stroki


def make_tabel_func(user, month, year, topic, topic_val, holiday, komandirovka, short_days_list, path, settings_list):

    wb = opx.Workbook()  # Создаём рабочую книгу
    ws = wb.active  # Запоминаем активный лист

    # Границы для форматирования ячеек
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    defalt_border = Border(left=Side(style=None),
                           right=Side(style=None),
                           top=Side(style=None),
                           bottom=Side(style=None))
    #  Размер шрифта 13
    ft = Font(name='Times New Roman',
              size=13,
              bold=False,
              italic=False,
              vertAlign=None,
              underline='none',
              strike=False,
              color='FF000000')
    # Номер месяца строкой, что бы вставлять 01, 02 и тд.
    if month < 10:
        monprint = '0' + str(month)
    else:
        monprint = str(month)

    # Словарь месяцев 1: "январь" и т.д.
    month_list = {1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель', 5: 'май', 6: 'июнь',
                  7: 'июль', 8: 'август', 9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декарбь'}

    # Количество активных тем
    topcount = 0
    for i, top in enumerate(topic):
        if top:
            ws[f'{chr(66+topcount)}4'].value = top
            topcount += 1

    # Делаем шапку
    ws.merge_cells(f'A1:{chr(66+topcount)}1')  # Объединяем ячейки для названия
    ws.merge_cells(f'A2:{chr(66+topcount)}2')  # Объединяем ячейки для названия
    ws.merge_cells(f'B3:{chr(66 + topcount - 1)}3')  # Объединяем ячейки для номера инвентарнго объекта
    ws.merge_cells(f'A3:A4')
    ws.merge_cells(f'{chr(66+topcount)}3:{chr(66+topcount)}4')

    # Добавляет в выбраную ячейку указанное значение
    ws['A1'].value = f'Табель учета затрат времени сотрудника ОВНТ {user}'
    ws['A2'].value = f'на «{month_list[month]}» {year} года по инвентарным объектам'
    ws['B3'].value = 'Номер инвентарного объекта'

    ws[f'{chr(66+topcount)}3'].value = 'Сумма'
    ws['A3'].value = 'Дата'

    w_day = workday(1, month, year, holiday, komandirovka)
    all_day = []

    for i, k in zip(topic, topic_val):
        if i:
            all_day.append(k)
    sum_total = [0 for _ in range(len(all_day) + 1)]

    tabel = fill_tabel(all_day, w_day, short_days_list)

    for ind, i in enumerate(tabel):
        day = [ind + 1]
        if type(i[0]) == int:
            qqq = [f'=SUM(B{day[0] + 4}:{chr(65 + topcount)}{day[0] + 4})']
            ws.append(day + i + qqq)
        else:
            ws.append(day + i + [' - '])
    for i in range(len(all_day) + 1):
        sum_total[i] = f'=SUM({chr(66 + i)}5:{chr(66 + i)}{len(tabel) + 4})'

    ws.append(['Итого'] + sum_total)

    rest_word = {0: 'ОТПУСК', 1: 'КОМАНДИРОВКА', 2: 'БОЛЬНИЧНЫЙ'}
    for rest in komandirovka:
        ws.merge_cells(f'B{rest[1] + 4}:{chr(66 + topcount)}{rest[2] + 4}')
        ws[f'B{rest[1] + 4}'].value = f'{rest_word[rest[0]]}'

    for row in ws:  # Проходим по всем строкам на листе
        for cell in row:  # Проходим по всем ячейкам в строке
            # Выравниваем ячейку по центру
            cell.alignment = opx.styles.Alignment(horizontal='center', vertical='center')
            cell.border = thin_border   # создаем границы клеткам
            cell.font = ft

    for col in range(65, 67 + topcount):
        for row in range(1, 3):
            ws[f'{chr(col)}{row}'].border = defalt_border

    for col in range(66, 67 + topcount):
        if col == 65 + topcount:
            ws.column_dimensions[f'{chr(col)}'].width = 20  # делаем шире столбец с Производством
        else:
            pass
            ws.column_dimensions[f'{chr(col)}'].width = 20 - 2 * topcount

    sum_width = 0
    for col in range(65, 68 + topcount):
        sum_width += ws.column_dimensions[f'{chr(col)}'].width

    ws.merge_cells(f'A41:{chr(66 + topcount)}41')

    if settings_list['current_nach']:
        cur_nach = 'nachalnik'
    else:
        cur_nach = 'io'
    ws['A41'].value = f'{settings_list[cur_nach][0]}' + \
                      (int(sum_width) - len(settings_list[cur_nach][0]) - len(settings_list[cur_nach][1])) * ' ' + \
                      f'{settings_list[cur_nach][1]}'
    ws['A41'].font = ft
    ws['A41'].alignment = opx.styles.Alignment(horizontal='center')

    os.chdir('..')
    path = os.getcwd()
    if not os.path.isdir(f'{year}'):
        os.mkdir(f'{year}')
    path += f'\\{year}'
    os.chdir(path)

    if not os.path.isdir(f'{monprint}.{month_list[month]}'):
        os.mkdir(f'{monprint}.{month_list[month]}')
    path += f'\\{monprint}.{month_list[month]}'
    os.chdir(path)
    wb.save(f'Табель_{user}_{month_list[month]}.{year}.xlsx')  # Сохраняем книгу
    path += f'.\\Табель_{user}_{month_list[month]}.{year}.xlsx'
    return [path, f'Табель_{user}_{month_list[month]}.{year}.xlsx']
